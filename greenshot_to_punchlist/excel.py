"""Workbook create/append for the punchlist.

One workbook per project. Each row records a screenshot-derived issue with
an embedded thumbnail in the Preview column.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from PIL import Image as PILImage


HEADERS = [
    "No.",
    "Date",
    "Project",
    "Discipline",
    "Title",
    "Description",
    "Priority",
    "Assigned To",
    "Status",
    "Screenshot",
    "Preview",
]

# Column letter -> width (Excel character units).
COLUMN_WIDTHS = {
    "A": 6,    # No.
    "B": 12,   # Date
    "C": 18,   # Project
    "D": 12,   # Discipline
    "E": 28,   # Title
    "F": 50,   # Description
    "G": 10,   # Priority
    "H": 18,   # Assigned To
    "I": 12,   # Status
    "J": 30,   # Screenshot filename
    "K": 22,   # Preview (thumbnail)
}

PREVIEW_COLUMN = "K"
DISCIPLINE_COLUMN = "D"
PRIORITY_COLUMN = "G"
STATUS_COLUMN = "I"

# Thumbnail sizing. Excel column width ~22 ≈ 155px; row height 60pt ≈ 80px.
# Keep the thumbnail inside that with a small margin.
THUMBNAIL_MAX_W = 145
THUMBNAIL_MAX_H = 75
# Render the embedded PNG at this multiple of the display size so the preview
# stays sharp when zoomed or viewed on a high-DPI screen. Excel scales the
# image down to the display dimensions set on the XLImage.
THUMBNAIL_SCALE = 4
ROW_HEIGHT_POINTS = 60

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_ALIGN = Alignment(vertical="center")
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
DATE_FORMAT = "yyyy-mm-dd"

_INVALID_FILENAME_CHARS = re.compile(r'[\\/:*?"<>|]+')


@dataclass
class Row:
    project: str
    discipline: str
    title: str
    description: str
    priority: str
    assigned_to: str
    status: str
    screenshot_path: Path


def sanitize_project_name(name: str) -> str:
    cleaned = _INVALID_FILENAME_CHARS.sub("_", name).strip().rstrip(".")
    if not cleaned or not cleaned.strip("._ "):
        return "punchlist"
    return cleaned


def workbook_path_for(project: str, workbook_dir: Path) -> Path:
    return workbook_dir / f"{sanitize_project_name(project)}.xlsx"


def _create_workbook(path: Path, disciplines: list[str], priorities: list[str]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Punchlist"
    ws.append(HEADERS)

    for col, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col].width = width

    for col_idx, _ in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    if disciplines:
        dv = DataValidation(
            type="list",
            formula1=f'"{",".join(disciplines)}"',
            allow_blank=True,
        )
        dv.add(f"{DISCIPLINE_COLUMN}2:{DISCIPLINE_COLUMN}1048576")
        ws.add_data_validation(dv)
    if priorities:
        dv = DataValidation(
            type="list",
            formula1=f'"{",".join(priorities)}"',
            allow_blank=True,
        )
        dv.add(f"{PRIORITY_COLUMN}2:{PRIORITY_COLUMN}1048576")
        ws.add_data_validation(dv)

    dv_status = DataValidation(
        type="list",
        formula1='"Open,In Progress,Resolved,Closed"',
        allow_blank=True,
    )
    dv_status.add(f"{STATUS_COLUMN}2:{STATUS_COLUMN}1048576")
    ws.add_data_validation(dv_status)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _next_number(ws) -> int:
    max_n = 0
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        value = row[0]
        if isinstance(value, int) and value > max_n:
            max_n = value
        elif isinstance(value, str) and value.strip().isdigit():
            n = int(value.strip())
            if n > max_n:
                max_n = n
    return max_n + 1


def _build_thumbnail(screenshot_path: Path) -> XLImage:
    pil_img = PILImage.open(screenshot_path)
    pil_img.thumbnail(
        (THUMBNAIL_MAX_W * THUMBNAIL_SCALE, THUMBNAIL_MAX_H * THUMBNAIL_SCALE),
        PILImage.LANCZOS,
    )

    if pil_img.mode not in ("RGB", "RGBA"):
        pil_img = pil_img.convert("RGB")

    buf = BytesIO()
    pil_img.save(buf, format="PNG")
    buf.seek(0)

    xl_img = XLImage(buf)
    xl_img.width = pil_img.width / THUMBNAIL_SCALE
    xl_img.height = pil_img.height / THUMBNAIL_SCALE
    return xl_img


def append_row(
    path: Path,
    row: Row,
    *,
    disciplines: list[str] | None = None,
    priorities: list[str] | None = None,
    today: date | None = None,
) -> int:
    """Append a row to the workbook at `path`. Creates it if missing.

    Returns the auto-assigned issue number.
    Raises PermissionError if the workbook is open in Excel.
    """
    if not path.exists():
        _create_workbook(path, disciplines or [], priorities or [])

    wb = load_workbook(path)
    ws = wb["Punchlist"] if "Punchlist" in wb.sheetnames else wb.active

    number = _next_number(ws)
    today = today or date.today()

    screenshot = row.screenshot_path
    filename = screenshot.name

    new_row = [
        number,
        today,
        row.project,
        row.discipline,
        row.title,
        row.description,
        row.priority,
        row.assigned_to,
        row.status,
        filename,
        None,
    ]
    ws.append(new_row)

    excel_row = ws.max_row
    ws.row_dimensions[excel_row].height = ROW_HEIGHT_POINTS

    date_cell = ws.cell(row=excel_row, column=2)
    date_cell.number_format = DATE_FORMAT

    title_cell = ws.cell(row=excel_row, column=5)
    title_cell.alignment = WRAP_ALIGN
    description_cell = ws.cell(row=excel_row, column=6)
    description_cell.alignment = WRAP_ALIGN

    try:
        thumbnail = _build_thumbnail(screenshot)
        thumbnail.anchor = f"{PREVIEW_COLUMN}{excel_row}"
        ws.add_image(thumbnail)
    except (FileNotFoundError, OSError, PILImage.UnidentifiedImageError):
        # Image missing or unreadable — still log the row, just no thumbnail.
        pass

    last_col = get_column_letter(len(HEADERS))
    ws.auto_filter.ref = f"A1:{last_col}{excel_row}"

    wb.save(path)
    return number
