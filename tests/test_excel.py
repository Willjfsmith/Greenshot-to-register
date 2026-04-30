from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

import pytest
from openpyxl import load_workbook
from PIL import Image as PILImage

from greenshot_to_punchlist import excel


@pytest.fixture
def screenshot(tmp_path: Path) -> Path:
    p = tmp_path / "shot.png"
    img = PILImage.new("RGB", (640, 360), color=(120, 140, 200))
    img.save(p, format="PNG")
    return p


def make_row(screenshot: Path, **overrides) -> excel.Row:
    defaults = dict(
        project="Project Alpha",
        discipline="ARCH",
        title="Door clashes column",
        description="The double door swing path overlaps the structural column on grid B3.",
        priority="High",
        assigned_to="J. Smith",
        status="Open",
        screenshot_path=screenshot,
    )
    defaults.update(overrides)
    return excel.Row(**defaults)


def test_creates_workbook_with_headers_and_appends_first_row(tmp_path, screenshot):
    path = excel.workbook_path_for("Project Alpha", tmp_path)
    assert not path.exists()

    n = excel.append_row(
        path,
        make_row(screenshot),
        disciplines=["ARCH", "STRUCT"],
        priorities=["High", "Low"],
        today=date(2026, 4, 29),
    )

    assert n == 1
    assert path.exists()

    wb = load_workbook(path)
    ws = wb["Punchlist"]
    headers = [c.value for c in ws[1]]
    assert headers == excel.HEADERS

    row2 = [c.value for c in ws[2]]
    assert row2[0] == 1
    # openpyxl reads dates back as datetime even when written as date.
    assert row2[1] == datetime(2026, 4, 29)
    assert row2[2] == "Project Alpha"
    assert row2[3] == "ARCH"
    assert row2[4] == "Door clashes column"
    assert row2[5].startswith("The double door")
    assert row2[6] == "High"
    assert row2[7] == "J. Smith"
    assert row2[8] == "Open"
    assert row2[9] == "shot.png"
    assert row2[10] is None  # Preview cell holds the image, not a value

    assert ws.cell(row=2, column=2).number_format == excel.DATE_FORMAT


def test_image_link_column_is_gone(tmp_path, screenshot):
    path = excel.workbook_path_for("Alpha", tmp_path)
    excel.append_row(path, make_row(screenshot))

    wb = load_workbook(path)
    ws = wb["Punchlist"]
    headers = [c.value for c in ws[1]]
    assert "Image Link" not in headers
    assert "Preview" in headers


def test_thumbnail_is_embedded(tmp_path, screenshot):
    path = excel.workbook_path_for("Alpha", tmp_path)
    excel.append_row(path, make_row(screenshot))

    wb = load_workbook(path)
    ws = wb["Punchlist"]
    assert len(ws._images) == 1
    img = ws._images[0]
    assert img.width <= excel.THUMBNAIL_MAX_W
    assert img.height <= excel.THUMBNAIL_MAX_H
    assert ws.row_dimensions[2].height == excel.ROW_HEIGHT_POINTS


def test_autofilter_extends_to_last_row(tmp_path, screenshot):
    path = excel.workbook_path_for("Alpha", tmp_path)
    excel.append_row(path, make_row(screenshot))
    excel.append_row(path, make_row(screenshot, title="Second"))

    wb = load_workbook(path)
    ws = wb["Punchlist"]
    assert ws.auto_filter.ref == "A1:K3"


def test_second_append_increments_number(tmp_path, screenshot):
    path = excel.workbook_path_for("Project Alpha", tmp_path)

    excel.append_row(path, make_row(screenshot), today=date(2026, 4, 29))
    n = excel.append_row(
        path, make_row(screenshot, title="Second issue"), today=date(2026, 4, 29)
    )

    assert n == 2

    wb = load_workbook(path)
    ws = wb["Punchlist"]
    assert ws.max_row == 3
    assert ws.cell(row=3, column=1).value == 2
    assert ws.cell(row=3, column=5).value == "Second issue"


def test_per_project_workbook_paths(tmp_path):
    p1 = excel.workbook_path_for("Project Alpha", tmp_path)
    p2 = excel.workbook_path_for("Project / Beta?", tmp_path)
    assert p1.name == "Project Alpha.xlsx"
    assert p2.name == "Project _ Beta_.xlsx"


def test_sanitize_handles_empty():
    assert excel.sanitize_project_name("") == "punchlist"
    assert excel.sanitize_project_name("///") == "punchlist"


def test_freeze_panes_and_header_style(tmp_path, screenshot):
    path = excel.workbook_path_for("Alpha", tmp_path)
    excel.append_row(path, make_row(screenshot))

    wb = load_workbook(path)
    ws = wb["Punchlist"]
    assert ws.freeze_panes == "A2"
    assert ws["A1"].font.bold is True


def test_missing_screenshot_still_logs_row(tmp_path):
    path = excel.workbook_path_for("Alpha", tmp_path)
    missing = tmp_path / "does-not-exist.png"
    row = excel.Row(
        project="Alpha",
        discipline="ARCH",
        title="No image",
        description="screenshot file missing",
        priority="Low",
        assigned_to="",
        status="Open",
        screenshot_path=missing,
    )
    n = excel.append_row(path, row)
    assert n == 1

    wb = load_workbook(path)
    ws = wb["Punchlist"]
    assert ws.cell(row=2, column=5).value == "No image"
    assert len(ws._images) == 0
